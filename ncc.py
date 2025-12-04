import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import load_workbook
import os

FILE_NAME = "Nominal Roll of B Cert Exam 2026.xlsx"

try:
    wb = load_workbook(FILE_NAME)
except:
    messagebox.showerror("Error", "Excel file not found in folder!")
    exit()

photo_path = ""

def upload_photo():
    global photo_path
    photo_path = filedialog.askopenfilename(
        title="Select Cadet Photo",
        filetypes=[("Image Files", "*.jpg *.jpeg *.png")]
    )
    if photo_path:
        messagebox.showinfo("Uploaded", "Photo selected successfully!")

def save_details():
    ser = ser_entry.get()
    regtl = regtl_entry.get()
    rank = rank_entry.get()
    name = name_entry.get()
    father = father_entry.get()
    dob = dob_entry.get()
    enroll = enroll_entry.get()
    discharge = discharge_entry.get()
    camp = camp_entry.get()
    attendance1 = att1_entry.get()
    attendance2 = att2_entry.get()

    if name.strip() == "":
        messagebox.showerror("Error", "Cadet Name is required!")
        return

    # Create separate sheet for each cadet
    if name not in wb.sheetnames:
        ws = wb.create_sheet(title=name)
        ws.append(["Field", "Details"])
    else:
        ws = wb[name]

    ws.append(["Ser No", ser])
    ws.append(["Regtl No", regtl])
    ws.append(["Rank", rank])
    ws.append(["Name of Cadet", name])
    ws.append(["Father Name", father])
    ws.append(["Date of Birth", dob])
    ws.append(["Date of Enrollment", enroll])
    ws.append(["Date of Discharge", discharge])
    ws.append(["Camp Details", camp])
    ws.append(["Parade Attendance (I Yr)", attendance1])
    ws.append(["Parade Attendance (II Yr)", attendance2])
    ws.append(["Photo Path", photo_path])
    ws.append([])

    wb.save(FILE_NAME)
    messagebox.showinfo("Success", "Cadet Details Added Successfully!")

    # Clear entries
    entries = [ser_entry, regtl_entry, rank_entry, name_entry, father_entry,
               dob_entry, enroll_entry, discharge_entry, camp_entry,
               att1_entry, att2_entry]
    for e in entries:
        e.delete(0, tk.END)

# ---------------- GUI -------------------

root = tk.Tk()
root.title("NCC Nominal Roll – Cadet Data Entry Form")
root.geometry("400x600")

tk.Label(root, text="Ser No").pack()
ser_entry = tk.Entry(root, width=30); ser_entry.pack()

tk.Label(root, text="Regtl No").pack()
regtl_entry = tk.Entry(root, width=30); regtl_entry.pack()

tk.Label(root, text="Rank").pack()
rank_entry = tk.Entry(root, width=30); rank_entry.pack()

tk.Label(root, text="Name of Cadet").pack()
name_entry = tk.Entry(root, width=30); name_entry.pack()

tk.Label(root, text="Father Name").pack()
father_entry = tk.Entry(root, width=30); father_entry.pack()

tk.Label(root, text="Date of Birth (DD-MM-YYYY)").pack()
dob_entry = tk.Entry(root, width=30); dob_entry.pack()

tk.Label(root, text="Date of Enrollment").pack()
enroll_entry = tk.Entry(root, width=30); enroll_entry.pack()

tk.Label(root, text="Date of Discharge").pack()
discharge_entry = tk.Entry(root, width=30); discharge_entry.pack()

tk.Label(root, text="Camp Details").pack()
camp_entry = tk.Entry(root, width=30); camp_entry.pack()

tk.Label(root, text="Parade Attendance % (I Yr)").pack()
att1_entry = tk.Entry(root, width=30); att1_entry.pack()

tk.Label(root, text="Parade Attendance % (II Yr)").pack()
att2_entry = tk.Entry(root, width=30); att2_entry.pack()

tk.Button(root, text="Upload Photo", command=upload_photo, bg="skyblue").pack(pady=10)

tk.Button(root, text="Save Cadet Details", command=save_details,
          bg="green", fg="white", width=25).pack(pady=20)

root.mainloop()
